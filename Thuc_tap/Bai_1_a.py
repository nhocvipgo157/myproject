import cv2 as cv
import openpyxl
from openpyxl import Workbook
import os


wb = Workbook()
ws = wb.active

img = cv.imread('/home/son/Myproject/Thuc_tap/Btap_1/cat.jpg')
(h,w,c) = img.shape
img_new = cv.resize(img,(int(h/2),int(w/2)))

file_path = "/home/son/Myproject/Thuc_tap/Btap_1/cat.jpg"
basename = os.path.basename(file_path)
name = os.path.splitext(basename)[0]
cv.imwrite("{}_{}x{}".format(name,int(h/2),int(w/2))+".jpg",img_new)

ws['A1'] = "Name"
ws['B1'] = "Size_old_img"
ws['C1'] = "Size_new_img"
ws.append([name,"{}x{}".format(h,w),"{}x{}".format(int(h/2),int(w/2))])
wb.save("/home/son/Myproject/Thuc_tap/Btap_1/Btap_1_a.xlsx")
